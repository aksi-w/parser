import openpyxl
from psycopg2 import connect, sql
from psycopg2.extras import execute_values


class Titular:
    def __init__(self, profile, beginning_year, fgos, program):
        self.profile = profile
        self.beginning_year = beginning_year
        self.fgos = fgos
        self.program = program

    def __repr__(self):
        return f"{self.profile} ({self.program})"


class Competention:
    def __init__(self, sub_index, index, description, type_):
        self.sub_index = sub_index
        self.index = index
        self.description = description
        self.type = type_

    def __repr__(self):
        return f"{self.index}: {self.description}"


class Plan:
    def __init__(self, count_in_plan, index, name, exam, midterm, kp, coursework, controlwork, midtermwithmark,
                 expert, factual, expert_hour, plan, controlwork_hour, aud, sr, control, preparation,
                 sem1, sem2, sem3, sem4, sem5, sem6, sem7, sem8, faculty_code, faculty_name):
        self.count_in_plan = count_in_plan
        self.index = index
        self.name = name
        self.exam = exam
        self.midterm = midterm
        self.kp = kp
        self.coursework = coursework
        self.controlwork = controlwork
        self.midtermwithmark = midtermwithmark
        self.expert = expert
        self.factual = factual
        self.expert_hour = expert_hour
        self.plan = plan
        self.controlwork_hour = controlwork_hour
        self.aud = aud
        self.sr = sr
        self.control = control
        self.preparation = preparation
        self.sem1 = sem1
        self.sem2 = sem2
        self.sem3 = sem3
        self.sem4 = sem4
        self.sem5 = sem5
        self.sem6 = sem6
        self.sem7 = sem7
        self.sem8 = sem8
        self.faculty_code = faculty_code
        self.faculty_name = faculty_name


PLAN_NAMES = {
    "Экзамен": "exam",
    "Зачет": "midterm",
    "Зачет с оц.": "midtermwithmark",
    "КП": "kp",
    "КР": "coursework",
    "Контр.": "controlwork",
}


def parse_titular(workbook):
    sheet = workbook["Titul"]
    profile = sheet.cell(row=30, column=4).value
    beginning_year_value = sheet.cell(row=40, column=23).value
    fgos = sheet.cell(row=42, column=23).value
    program = sheet.cell(row=29, column=4).value

    if not profile:
        raise ValueError("Profile value is missing in the Titul sheet")
    if beginning_year_value is None:
        raise ValueError("Beginning year value is missing in the Titul sheet")
    try:
        beginning_year = int(beginning_year_value)
    except ValueError:
        raise ValueError(f"Invalid beginning year value: {beginning_year_value}")
    if not fgos:
        raise ValueError("FGOS value is missing in the Titul sheet")
    if not program:
        raise ValueError("Program value is missing in the Titul sheet")

    return Titular(profile, beginning_year, fgos, program)


def parse_competentions(workbook):
    sheet = workbook["Comp"]
    competention_list = []
    for i in range(2, sheet.max_row + 1):
        sub_index = sheet.cell(row=i, column=2).value or ""
        index = sheet.cell(row=i, column=3).value or ""
        description = sheet.cell(row=i, column=5).value or ""
        type_ = sheet.cell(row=i, column=6).value or ""
        if description:
            competention_list.append(Competention(sub_index, index, description, type_))
    return competention_list


def parse_plans(workbook):
    sheet = workbook["PlanSvod"]
    plans = []

    for i in range(6, sheet.max_row + 1):
        row = sheet[i]
        if row[0].value in ("+", "-"):
            plans.append(
                Plan(
                    count_in_plan=row[0].value,
                    index=f"{row[2].value or ''} {row[3].value or ''}".strip(),
                    name=row[4].value or "",
                    exam=row[6].value or "",
                    midterm=row[8].value or "",
                    midtermwithmark=row[10].value or "",
                    kp=row[12].value or "",
                    coursework=row[18].value or "",
                    controlwork=row[10].value or "",
                    expert=row[18].value or "",
                    factual=row[20].value or "",
                    expert_hour=row[22].value or "",
                    plan=row[24].value or "",
                    controlwork_hour=row[26].value or "",
                    aud=row[28].value or "",
                    sr=row[30].value or "",
                    control=row[32].value or "",
                    preparation=row[34].value or "",
                    sem1=row[36].value or "",
                    sem2=row[38].value or "",
                    sem3=row[40].value or "",
                    sem4=row[42].value or "",
                    sem5=row[44].value or "",
                    sem6=row[46].value or "",
                    sem7=row[48].value or "",
                    sem8=row[50].value or "",
                    faculty_code=row[52].value or "",
                    faculty_name=row[54].value or "",
                )
            )
    return plans


def create_tables(conn):
    with conn.cursor() as cur:
        cur.execute("DROP TABLE IF EXISTS titulars, competentions, plans CASCADE")
        cur.execute("""
            CREATE TABLE titulars (
                profile VARCHAR,
                beginning_year BIGINT,
                fgos VARCHAR,
                program VARCHAR
            );
        """)
        cur.execute("""
            CREATE TABLE competentions (
                sub_index VARCHAR,
                index VARCHAR,
                description VARCHAR,
                type VARCHAR
            );
        """)
        cur.execute("""
            CREATE TABLE plans (
                count_in_plan VARCHAR,
                index VARCHAR,
                name VARCHAR,
                exam VARCHAR,
                midterm VARCHAR,
                kp VARCHAR,
                coursework VARCHAR,
                controlwork VARCHAR,
                midtermwithmark VARCHAR,
                expert VARCHAR,
                factual VARCHAR,
                expert_hour VARCHAR,
                plan VARCHAR,
                controlwork_hour VARCHAR,
                aud VARCHAR,
                sr VARCHAR,
                control VARCHAR,
                preparation VARCHAR,
                sem1 VARCHAR,
                sem2 VARCHAR,
                sem3 VARCHAR,
                sem4 VARCHAR,
                sem5 VARCHAR,
                sem6 VARCHAR,
                sem7 VARCHAR,
                sem8 VARCHAR,
                faculty_code VARCHAR,
                faculty_name VARCHAR
            );
        """)


def insert_data(conn, titular, competentions, plans):
    with conn.cursor() as cur:
        cur.execute(
            "INSERT INTO titulars (profile, beginning_year, fgos, program) VALUES (%s, %s, %s, %s)",
            (titular.profile, titular.beginning_year, titular.fgos, titular.program),
        )

        competention_values = [
            (comp.sub_index, comp.index, comp.description, comp.type) for comp in competentions
        ]
        execute_values(
            cur,
            "INSERT INTO competentions (sub_index, index, description, type) VALUES %s",
            competention_values,
        )

        plan_values = [
            (
                plan.count_in_plan,
                plan.index or "",
                plan.name or "",
                plan.exam or "",
                plan.midterm or "",
                plan.kp or "",
                plan.coursework or "",
                plan.controlwork or "",
                plan.midtermwithmark or "",
                plan.expert or "",
                plan.factual or "",
                plan.expert_hour or "",
                plan.plan or "",
                plan.controlwork_hour or "",
                plan.aud or "",
                plan.sr or "",
                plan.control or "",
                plan.preparation or "",
                plan.sem1 or "",
                plan.sem2 or "",
                plan.sem3 or "",
                plan.sem4 or "",
                plan.sem5 or "",
                plan.sem6 or "",
                plan.sem7 or "",
                plan.sem8 or "",
                plan.faculty_code or "",
                plan.faculty_name or "",
            )
            for plan in plans
        ]

        execute_values(
            cur,
            "INSERT INTO plans (count_in_plan, index, name, exam, midterm, kp, coursework, "
            "controlwork, midtermwithmark, expert, factual, expert_hour, plan, controlwork_hour, "
            "aud, sr, control, preparation, sem1,sem2,sem3,sem4,sem5,sem6,sem7,sem8, faculty_code, faculty_name) VALUES %s",
            plan_values,
        )


def start_parsing(filename, connection_string):
    conn = connect(connection_string)
    workbook = openpyxl.load_workbook(filename)

    titular = parse_titular(workbook)
    competentions = parse_competentions(workbook)
    plans = parse_plans(workbook)

    create_tables(conn)
    insert_data(conn, titular, competentions, plans)

    conn.commit()
    conn.close()
